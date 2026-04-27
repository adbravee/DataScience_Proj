"""
===================================================================
        SMART CLASSROOM ATTENDANCE SYSTEM                         
        Powered by ArcFace CNN + RetinaFace                       
        Production-Ready Build: RBAC, Continuous Learning, Exports            
===================================================================
"""

import os
import pickle
import warnings
from io import BytesIO
from datetime import datetime, date, timezone, timedelta

import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image
from scipy.spatial.distance import cosine

warnings.filterwarnings("ignore")

# Define Indian Standard Time (UTC + 5:30)
IST = timezone(timedelta(hours=5, minutes=30))

# -----------------------------------------------------------------
#  IMPORT GUARDS
# -----------------------------------------------------------------
def check_imports():
    errors = []
    try:
        import cv2
    except ImportError:
        errors.append("cv2 (OpenCV)")
    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl (Excel support)")
    try:
        import retinaface
    except ImportError:
        errors.append("retina-face (Crucial for group detection)")
    
    if errors:
        st.error(f"Missing Python packages: {', '.join(errors)}\nFix: pip install opencv-python openpyxl retina-face")
        st.stop()

check_imports()

import cv2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------
#  CONFIGURATION
# -----------------------------------------------------------------
class Config:
    # --- SECURITY ---
    ADMIN_PASSWORD    = "Bologna"  # Master password for scanning and data management
    
    # --- MODEL ---
    MODEL_NAME        = "ArcFace"
    DETECTOR_BACKEND  = "retinaface"
    DEFAULT_THRESHOLD = 0.60
    
    # --- CONTINUOUS LEARNING ---
    AUTO_UPDATE_THRESHOLD = 0.82 
    MAX_RECORDS_PER_IDENTITY = 15
    
    # --- PATHS ---
    DATA_DIR          = "data"
    EMBEDDINGS_FILE   = "data/student_embeddings.pkl"
    ATTENDANCE_FILE   = "data/attendance.xlsx"
    STUDENT_IMG_DIR   = "data/student_images"


# -----------------------------------------------------------------
#  BOOTSTRAP
# -----------------------------------------------------------------
def init_directories() -> None:
    for path in [Config.DATA_DIR, Config.STUDENT_IMG_DIR]:
        os.makedirs(path, exist_ok=True)


# -----------------------------------------------------------------
#  MODEL LOADER
# -----------------------------------------------------------------
@st.cache_resource(show_spinner="Loading AI Models (ArcFace & RetinaFace)...")
def load_deepface():
    from deepface import DeepFace
    dummy = np.zeros((160, 160, 3), dtype=np.uint8)
    try:
        DeepFace.represent(
            img_path=dummy,
            model_name=Config.MODEL_NAME,
            detector_backend="skip",
            enforce_detection=False,
        )
    except Exception:
        pass
    return DeepFace


# -----------------------------------------------------------------
#  EMBEDDING EXTRACTION & DETECTION (CNN)
# -----------------------------------------------------------------
def l2_normalize(vec: np.ndarray) -> np.ndarray:
    return vec / (np.linalg.norm(vec) + 1e-10)

def extract_embedding_single(img_rgb: np.ndarray, DeepFace) -> tuple:
    img_bgr = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2BGR)
    try:
        results = DeepFace.represent(
            img_path=img_bgr,
            model_name=Config.MODEL_NAME,
            detector_backend=Config.DETECTOR_BACKEND,
            enforce_detection=True,
            align=True, 
        )
        if results:
            emb = l2_normalize(np.array(results[0]["embedding"]))
            return emb, results[0].get("facial_area", {})
    except Exception as e:
        st.debug(f"Registration Error: {e}")
    return None, {}

def extract_all_faces(img_rgb: np.ndarray, DeepFace) -> list:
    if img_rgb is None or img_rgb.size == 0:
        return []
    
    img_bgr = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2BGR)
    extracted_faces = []
    
    try:
        results = DeepFace.represent(
            img_path=img_bgr,
            model_name=Config.MODEL_NAME,
            detector_backend=Config.DETECTOR_BACKEND,
            enforce_detection=False,
            align=True,
        )
        
        for r in results:
            if "embedding" not in r:
                continue
            
            confidence = r.get("face_confidence", 1.0)
            
            # Accuracy Filter: Reject background artifacts
            if confidence < 0.90:
                continue
                
            facial_area = r.get("facial_area", {})
            if facial_area:
                if min(facial_area.get("w", 0), facial_area.get("h", 0)) < 15:
                    continue
                    
            emb = l2_normalize(np.array(r["embedding"]))
            extracted_faces.append({
                "embedding": emb,
                "facial_area": facial_area,
                "confidence": confidence,
            })
            
        return extracted_faces
        
    except Exception as e:
        st.debug(f"Group Detection Error: {e}")
        return []


# -----------------------------------------------------------------
#  DATABASE & CONTINUOUS LEARNING MANAGEMENT
# -----------------------------------------------------------------
def load_embeddings() -> dict:
    if os.path.exists(Config.EMBEDDINGS_FILE):
        with open(Config.EMBEDDINGS_FILE, "rb") as f:
            return pickle.load(f)
    return {}

def save_embeddings(db: dict) -> None:
    with open(Config.EMBEDDINGS_FILE, "wb") as f:
        pickle.dump(db, f)

def add_student(identifier: str, embedding: np.ndarray, thumb_rgb: np.ndarray = None) -> None:
    db = load_embeddings()
    if identifier not in db:
        db[identifier] = {"embeddings": [], "thumbnail": None}
    db[identifier]["embeddings"].append(embedding)
    if thumb_rgb is not None:
        db[identifier]["thumbnail"] = cv2.resize(thumb_rgb, (80, 80))
    save_embeddings(db)

def delete_student(identifier: str) -> None:
    db = load_embeddings()
    db.pop(identifier, None)
    save_embeddings(db)

def batch_update_embeddings(updates: dict) -> int:
    if not updates:
        return 0
        
    db = load_embeddings()
    update_count = 0
    
    for identifier, new_emb in updates.items():
        if identifier in db:
            db[identifier]["embeddings"].append(new_emb)
            
            # Enforce memory limit
            if len(db[identifier]["embeddings"]) > Config.MAX_RECORDS_PER_IDENTITY:
                original_reg = db[identifier]["embeddings"][0]
                recent_recs = db[identifier]["embeddings"][-(Config.MAX_RECORDS_PER_IDENTITY - 1):]
                db[identifier]["embeddings"] = [original_reg] + recent_recs
                
            update_count += 1
            
    if update_count > 0:
        save_embeddings(db)
        
    return update_count


# -----------------------------------------------------------------
#  FACE MATCHING
# -----------------------------------------------------------------
def match_face(query_emb: np.ndarray, db: dict, threshold: float) -> tuple:
    best_identifier = "Unknown"
    best_score = -1.0

    for identifier, data in db.items():
        scores = [1.0 - cosine(query_emb, e) for e in data["embeddings"]]
        score = max(scores) if scores else 0.0
        if score > best_score:
            best_score = score
            best_identifier = identifier

    if best_score < threshold:
        return "Unknown", best_score
    return best_identifier, best_score


# -----------------------------------------------------------------
#  ATTENDANCE LOGIC
# -----------------------------------------------------------------
def load_attendance() -> pd.DataFrame:
    if os.path.exists(Config.ATTENDANCE_FILE):
        try:
            return pd.read_excel(Config.ATTENDANCE_FILE)
        except Exception:
            pass
    return pd.DataFrame(columns=["Roll Number", "Name", "Date", "Time", "Status"])

def save_attendance(df: pd.DataFrame) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hdr_font = Font(name="Calibri", color="FFFFFF", bold=True, size=12)
    hdr_border = Border(bottom=Side(style="medium", color="4A90D9"))

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = hdr_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    even_fill = PatternFill(start_color="EBF3FC", end_color="EBF3FC", fill_type="solid")
    odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=11)

    for r_idx, row in df.iterrows():
        fill = even_fill if r_idx % 2 == 0 else odd_fill
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx + 2, column=c_idx, value=value)
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.iter_cols():
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 6

    ws.row_dimensions[1].height = 22
    wb.save(Config.ATTENDANCE_FILE)

def mark_attendance(roll_number: str, name: str, date_str: str) -> tuple[bool, str]:
    now = datetime.now(IST).strftime("%H:%M:%S")
    df  = load_attendance()

    if not df.empty:
        dup = df[(df["Roll Number"] == roll_number) & (df["Date"] == date_str)]
        if not dup.empty:
            return False, "Already marked"

    new_row = pd.DataFrame([{
        "Roll Number": roll_number,
        "Name":        name,
        "Date":        date_str,
        "Time":        now,
        "Status":      "Present",
    }])
    df = pd.concat([df, new_row], ignore_index=True)
    save_attendance(df)
    return True, f"Marked at {now}"


# -----------------------------------------------------------------
#  IMAGE ANNOTATION
# -----------------------------------------------------------------
def annotate_image(img_rgb: np.ndarray, detections: list) -> np.ndarray:
    img = img_rgb.copy()
    h, w = img.shape[:2]

    for det in detections:
        identifier = det["identifier"]
        score = det["similarity"]
        area  = det.get("facial_area", {})

        x = area.get("x", 0); y = area.get("y", 0)
        bw = area.get("w", 0); bh = area.get("h", 0)

        if not (bw > 0 and bh > 0):
            continue

        is_recognized = identifier != "Unknown"
        if is_recognized:
            if score >= Config.AUTO_UPDATE_THRESHOLD:
                colour = (46, 213, 115) 
            elif score >= Config.DEFAULT_THRESHOLD:
                colour = (75, 192, 192) 
            else:
                colour = (255, 193, 7)  
        else:
            colour = (255, 71, 87)      

        thick = max(2, min(w, h) // 200)
        cv2.rectangle(img, (x, y), (x + bw, y + bh), colour, thick)

        corner_len = max(10, bw // 6)
        for cx, cy, dx, dy in [
            (x, y, 1, 1), (x + bw, y, -1, 1),
            (x, y + bh, 1, -1), (x + bw, y + bh, -1, -1),
        ]:
            cv2.line(img, (cx, cy), (cx + dx * corner_len, cy), colour, thick + 2)
            cv2.line(img, (cx, cy), (cx, cy + dy * corner_len), colour, thick + 2)

        label = identifier if is_recognized else "Unknown"
        label = f"{label} ({score:.2f})"
        
        font     = cv2.FONT_HERSHEY_DUPLEX
        font_scale = max(0.4, min(w, h) / 1000)
        (tw, th), baseline = cv2.getTextSize(label, font, font_scale, 1)
        pad = 6
        label_y  = y - pad if y - th - pad * 3 >= 0 else y + bh + th + pad * 3

        cv2.rectangle(
            img,
            (x, label_y - th - pad),
            (x + tw + pad * 2, label_y + baseline + pad),
            colour, -1,
        )
        cv2.putText(
            img, label,
            (x + pad, label_y),
            font, font_scale, (255, 255, 255), 1, cv2.LINE_AA,
        )
        
        bar_height = max(3, bh // 25)
        bar_width = bw
        bar_y = y + bh + 2
        
        cv2.rectangle(img, (x, bar_y), (x + bar_width, bar_y + bar_height), (100, 100, 100), -1)
        filled_width = int(bar_width * max(0, min(1, score)))
        cv2.rectangle(img, (x, bar_y), (x + filled_width, bar_y + bar_height), colour, -1)
    
    return img


# -----------------------------------------------------------------
#  STREAMLIT UI & CSS
# -----------------------------------------------------------------
CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

:root {
    --bg:        #0E1117;
    --surface:   #161B22;
    --surface2:  #21262D;
    --border:    #30363D;
    --accent:    #58A6FF;
    --text:      #C9D1D9;
    --muted:     #8B949E;
    --font:      'Inter', sans-serif;
}

html, body, [class*="css"] {
    font-family: var(--font) !important;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}

section[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border);
}

.hero {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 24px;
    margin-bottom: 24px;
}
.hero h1 {
    font-size: 1.6rem; font-weight: 600;
    color: #FFFFFF !important;
    margin: 0 0 8px 0;
}
.hero p {
    color: var(--muted) !important;
    font-size: 0.95rem; margin: 0;
}

.section-title {
    font-size: 1.1rem; font-weight: 600;
    color: #FFFFFF;
    border-bottom: 1px solid var(--border);
    padding-bottom: 8px; margin-bottom: 16px;
}

div.stButton > button {
    background: var(--surface2) !important;
    color: #FFFFFF !important;
    border: 1px solid var(--border);
    border-radius: 6px;
    font-weight: 500;
}
div.stButton > button:hover { 
    background: var(--border) !important; 
}

div[data-testid="stFileUploader"] {
    border: 1px dashed var(--border) !important;
    border-radius: 6px;
    background: var(--surface);
}

.lock-screen {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 40px 20px;
    text-align: center;
    color: var(--muted);
}
.lock-screen h3 {
    color: #FFFFFF;
    margin-bottom: 10px;
}

/* Explicit Post-Scan Grid Layout */
.result-header {
    display: grid; 
    grid-template-columns: 80px 1fr 120px 100px 60px; 
    gap: 10px; 
    padding: 0 14px 8px 14px; 
    font-weight: 600; 
    color: var(--muted); 
    font-size: 0.85rem; 
    border-bottom: 1px solid var(--border); 
    margin-bottom: 8px;
}
.result-row {
    display: grid; 
    grid-template-columns: 80px 1fr 120px 100px 60px; 
    gap: 10px; 
    align-items: center;
    background: var(--surface); 
    border: 1px solid var(--border);
    border-radius: 6px; 
    padding: 10px 14px; 
    margin-bottom: 8px;
}

.chip-present  { background: rgba(63,185,80,.1);  color: #3FB950; border: 1px solid rgba(63,185,80,.3);  border-radius: 12px; padding: 2px 8px; font-size:0.75rem; font-weight:500; text-align: center; }
.chip-updated  { background: rgba(88,166,255,.1); color: #58A6FF; border: 1px solid rgba(88,166,255,.3); border-radius: 12px; padding: 2px 8px; font-size:0.75rem; font-weight:500; text-align: center;}
.chip-already  { background: rgba(210,153,34,.1); color: #D29922; border: 1px solid rgba(210,153,34,.3); border-radius: 12px; padding: 2px 8px; font-size:0.75rem; font-weight:500; text-align: center;}
.chip-unknown  { background: rgba(255,123,114,.1);color: #FF7B72; border: 1px solid rgba(255,123,114,.3);border-radius: 12px; padding: 2px 8px; font-size:0.75rem; font-weight:500; text-align: center;}

.conf-bar-wrap { background: var(--bg); border-radius: 10px; height:4px; width: 100%; }
.conf-bar      { height:4px; border-radius:10px; background: var(--accent); }
</style>
"""

def hero_banner():
    st.markdown("""
    <div class="hero">
      <h1>Smart Attendance System</h1>
      <p>Automated face-recognition utilizing ArcFace CNN and RetinaFace Group Detection.</p>
    </div>
    """, unsafe_allow_html=True)


# -----------------------------------------------------------------
#  PAGE: REGISTER
# -----------------------------------------------------------------
def page_register(DeepFace, is_admin):
    st.markdown('<p class="section-title">Student Registration</p>', unsafe_allow_html=True)
    
    c1, c2 = st.columns([1.1, 1], gap="large")

    with c1:
        rc1, rc2 = st.columns([1, 2])
        with rc1:
            roll_number = st.text_input("Roll Number", placeholder="e.g. 101")
        with rc2:
            name = st.text_input("Full Name", placeholder="Enter student name")
            
        mode = st.radio("Upload Method", [ "Multiple photos"], horizontal=True)
        
        can_proceed = False
        if mode == "Single photo":
            files = st.file_uploader("Select Image", type=["jpg", "jpeg", "png"])
            files = [files] if files else []
            if files:
                can_proceed = True
        else:
            files = st.file_uploader("Select Images (Minimum 3 required)", type=["jpg", "jpeg", "png"], accept_multiple_files=True)
            if files:
                if len(files) < 3:
                    st.warning("System Policy: Please upload at least 3 images from different angles to ensure high recognition accuracy.")
                    can_proceed = False
                else:
                    can_proceed = True

        if can_proceed and name.strip() and roll_number.strip():
            if st.button("Register Student Data", use_container_width=True):
                identifier = f"{roll_number.strip()} - {name.strip()}"
                
                bar = st.progress(0, text="Processing...")
                ok = 0
                for i, f in enumerate(files):
                    img  = np.array(Image.open(f).convert("RGB"))
                    with st.spinner(f"Encoding facial data {i+1}/{len(files)}..."):
                        emb, _ = extract_embedding_single(img, DeepFace)
                    if emb is not None:
                        add_student(identifier, emb, img)
                        ok += 1
                    bar.progress((i + 1) / len(files), text=f"Processed {i+1}/{len(files)}")

                if ok:
                    st.success(f"Registration Complete: {identifier} registered with {ok} data points.")
                else:
                    st.error("Registration Failed: No valid faces detected in the provided images.")
        elif files and (not name.strip() or not roll_number.strip()):
            st.warning("Both Roll Number and Student Name fields are strictly required.")

    with c2:
        if is_admin:
            st.markdown('<p class="section-title">Database Overview</p>', unsafe_allow_html=True)
            db = load_embeddings()
            if not db:
                st.info("System database is currently empty.")
            else:
                for identifier, data in db.items():
                    with st.expander(f"{identifier} ({len(data['embeddings'])} records)"):
                        cols = st.columns([1, 2])
                        if data.get("thumbnail") is not None:
                            cols[0].image(data["thumbnail"], width=100) 
                        
                        try:
                            roll, disp_name = identifier.split(" - ", 1)
                        except ValueError:
                            roll, disp_name = "N/A", identifier
                            
                        cols[1].markdown(f"**Roll No:** {roll}")
                        cols[1].markdown(f"**Name:** {disp_name}")
                        cols[1].markdown(f"**Data Points:** {len(data['embeddings'])}")
                        
                        if cols[1].button("Remove Record", key=f"del_{identifier}"):
                            delete_student(identifier)
                            st.rerun()
        else:
            st.markdown('<p class="section-title">Database Overview</p>', unsafe_allow_html=True)
            st.info("Database Overview is restricted to Administrators. Please enter the Access Key in the sidebar to view records.")


# -----------------------------------------------------------------
#  PAGE: MARK ATTENDANCE
# -----------------------------------------------------------------
def page_attendance(DeepFace, threshold, is_admin):
    st.markdown('<p class="section-title">Classroom Scan</p>', unsafe_allow_html=True)
    
    if not is_admin:
        st.markdown("""
        <div class="lock-screen">
            <h3>Administrator Access Required</h3>
            <p>Classroom scanning and attendance logging are locked. Please enter the Access Key in the sidebar menu to unlock this module.</p>
        </div>
        """, unsafe_allow_html=True)
        return

    db = load_embeddings()
    if not db:
        st.warning("Database empty. Please register students prior to scanning.")
        return

    c1, c2 = st.columns([1, 1.1], gap="large")

    with c1:
        upl = st.file_uploader("Select Classroom Image", type=["jpg", "jpeg", "png"])
        att_date = st.date_input("Date of Record", value=datetime.now(IST).date())

        if upl:
            img_rgb = np.array(Image.open(upl).convert("RGB"))
            st.image(img_rgb, caption="Input Source", use_container_width=True)

        if upl and st.button("Execute Scan & Log Attendance", use_container_width=True):
            with st.spinner("Executing RetinaFace detection matrix..."):
                faces = extract_all_faces(img_rgb, DeepFace)

            if not faces:
                st.error("Detection Failed: Zero faces located in the source image.")
                return

            date_str    = att_date.strftime("%Y-%m-%d")
            detections  = []
            results_log = []
            pending_updates = {}

            for face in faces:
                identifier, score = match_face(face["embedding"], db, threshold)
                detections.append({"identifier": identifier, "similarity": score, "facial_area": face["facial_area"]})

                if identifier != "Unknown":
                    try:
                        roll_no, stu_name = identifier.split(" - ", 1)
                    except ValueError:
                        roll_no, stu_name = "N/A", identifier
                        
                    marked, msg = mark_attendance(roll_no, stu_name, date_str)
                    
                    status = "present" if marked else "already"
                    
                    if score >= Config.AUTO_UPDATE_THRESHOLD:
                        pending_updates[identifier] = face["embedding"]
                        status = "updated" if marked else "already_updated"

                    results_log.append({
                        "roll":   roll_no,
                        "name":   stu_name,
                        "score":  score,
                        "status": status,
                    })
                else:
                    results_log.append({
                        "roll": "N/A", 
                        "name": "Unknown", 
                        "score": score, 
                        "status": "unknown"
                    })

            if pending_updates:
                updated_count = batch_update_embeddings(pending_updates)
                st.success(f"Scan Complete: Located {len(faces)} subjects. Model learned {updated_count} new variations.")
            else:
                st.success(f"Scan Complete: Located {len(faces)} subjects.")

            st.session_state["last_annotated"] = annotate_image(img_rgb, detections)
            st.session_state["last_results"]   = results_log

    with c2:
        if "last_annotated" in st.session_state:
            st.image(st.session_state["last_annotated"], caption="Processed Output", use_container_width=True)

        if "last_results" in st.session_state:
            results_log = st.session_state["last_results"]
            st.markdown('<p class="section-title">Match Results</p>', unsafe_allow_html=True)

            # Explicit Header for Columns
            st.markdown("""
            <div class="result-header">
                <span>Roll No</span>
                <span>Name</span>
                <span>Status</span>
                <span>Match Conf.</span>
                <span>Score</span>
            </div>
            """, unsafe_allow_html=True)

            for r in results_log:
                status_map = {
                    "present": ("chip-present", "Logged"),
                    "updated": ("chip-updated", "Logged + Learned"),
                    "already": ("chip-already", "Duplicate"),
                    "already_updated": ("chip-updated", "Dup + Learned"),
                    "unknown": ("chip-unknown", "Unknown")
                }
                
                chip_cls, chip_lbl = status_map.get(r["status"], ("chip-unknown", "Unknown"))
                bar_w    = int(r["score"] * 100)
                bar_col  = "#3FB950" if "unknown" not in r["status"] else "#FF7B72"
                
                if "updated" in r["status"]:
                    bar_col = "#58A6FF"
                
                roll_disp = r['roll'] if r['status'] != "unknown" else "—"
                name_disp = r['name'] if r['status'] != "unknown" else "Unidentified Subject"
                
                st.markdown(f"""
                <div class="result-row">
                  <span style="font-weight:600; font-family: monospace;">{roll_disp}</span>
                  <span style="font-weight:500;">{name_disp}</span>
                  <span class="{chip_cls}">{chip_lbl}</span>
                  <div class="conf-bar-wrap">
                    <div class="conf-bar" style="width:{bar_w}%;background:{bar_col}"></div>
                  </div>
                  <span style="font-size:0.85rem;color:var(--muted);text-align:right;">{r['score']:.2f}</span>
                </div>
                """, unsafe_allow_html=True)

            present = sum(1 for r in results_log if "present" in r["status"] or "updated" in r["status"])
            already = sum(1 for r in results_log if "already" in r["status"])
            unknown = sum(1 for r in results_log if r["status"] == "unknown")
            learned = sum(1 for r in results_log if "updated" in r["status"])
            
            st.markdown("<br>", unsafe_allow_html=True)
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Newly Marked", present)
            m2.metric("Duplicates", already)
            m3.metric("Unidentified", unknown)
            m4.metric("Profiles Learned", learned)


# -----------------------------------------------------------------
#  PAGE: VIEW RECORDS
# -----------------------------------------------------------------
def page_records():
    st.markdown('<p class="section-title">System Logs</p>', unsafe_allow_html=True)

    df = load_attendance()
    if df.empty:
        st.info("No records present in the database.")
        return

    fc1, fc2, fc3 = st.columns([1, 1, 1])
    with fc1:
        dates = ["All"] + sorted(df["Date"].unique(), reverse=True)
        sel_date = st.selectbox("Date Filter", dates)
    with fc2:
        students = ["All"] + sorted(df["Name"].unique())
        sel_stu  = st.selectbox("Name Filter", students)

    filt = df.copy()
    if sel_date != "All":
        filt = filt[filt["Date"] == sel_date]
    if sel_stu != "All":
        filt = filt[filt["Name"] == sel_stu]

    with fc3:
        st.metric("Total Displayed", len(filt))

    st.dataframe(filt, use_container_width=True, hide_index=True)

    st.markdown('<p class="section-title" style="margin-top:24px">Aggregate Data</p>', unsafe_allow_html=True)

    total_days = df["Date"].nunique()
    summary = (
        df.groupby(["Roll Number", "Name"])
        .agg(Days_Present=("Date", "nunique"), First_Seen=("Date", "min"))
        .reset_index()
    )
    summary["Attendance_%"] = (summary["Days_Present"] / total_days * 100).round(1).astype(str) + "%"
    st.dataframe(summary, use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)
    d1, d2 = st.columns(2)
    
    # Export CSV dynamically based on filter
    with d1:
        csv = filt.to_csv(index=False)
        date_label = sel_date if sel_date != "All" else "All_Dates"
        st.download_button("Export Filtered CSV", csv, f"Attendance_{date_label}.csv", "text/csv", use_container_width=True)
        
    # Export Excel dynamically based on filter using BytesIO
    with d2:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            filt.to_excel(writer, index=False, sheet_name='Attendance Logs')
        excel_data = output.getvalue()
        
        st.download_button(
            "Export Filtered Excel", 
            excel_data, 
            f"Attendance_{date_label}.xlsx", 
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            use_container_width=True
        )


# -----------------------------------------------------------------
#  PAGE: SETTINGS
# -----------------------------------------------------------------
def page_settings(threshold, is_admin):
    st.markdown('<p class="section-title">Configuration</p>', unsafe_allow_html=True)

    c1, c2 = st.columns(2, gap="large")
    with c1:
        st.subheader("Active Parameters")
        st.markdown(f"""
| Property | Value |
|-----------|-------|
| Neural Net | `{Config.MODEL_NAME}` |
| Detection CNN | `{Config.DETECTOR_BACKEND}` |
| Distance Metric | Cosine |
| Active Cutoff | `{threshold:.2f}` |
| Administrator Mode | `{"Active" if is_admin else "Disabled"}` |
| Continuous Learning | `Active` |
| Learning Threshold | `{Config.AUTO_UPDATE_THRESHOLD:.2f}` |
        """)

    with c2:
        st.subheader("Data Wipes")
        
        if is_admin:
            if st.button("Purge Attendance Logs", use_container_width=True):
                if os.path.exists(Config.ATTENDANCE_FILE):
                    os.remove(Config.ATTENDANCE_FILE)
                st.success("Logs purged successfully.")
                st.rerun()

            if st.button("Purge Registered Identities", use_container_width=True):
                if os.path.exists(Config.EMBEDDINGS_FILE):
                    os.remove(Config.EMBEDDINGS_FILE)
                st.success("Identities purged successfully.")
                st.rerun()
        else:
            st.info("System modification is restricted. Please enter the Admin Access Key in the sidebar to perform data wipes.")


# -----------------------------------------------------------------
#  MAIN LOOP
# -----------------------------------------------------------------
def main():
    st.set_page_config(
        page_title="Smart Attendance Portal",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
    init_directories()
    
    if "is_admin" not in st.session_state:
        st.session_state["is_admin"] = False

    with st.sidebar:
        st.markdown("### Navigation")
        
        nav_options = ["Student Registration", "Classroom Scan", "Configuration"]
        if st.session_state["is_admin"]:
            nav_options.insert(2, "System Logs")
            
        page = st.radio(
            "Menu",
            nav_options,
            label_visibility="collapsed",
        )

        st.markdown("---")
        db   = load_embeddings()
        att  = load_attendance()
        today_str = datetime.now(IST).strftime("%Y-%m-%d")
        today_cnt = len(att[att["Date"] == today_str]) if not att.empty else 0

        st.metric("Identities Enrolled", len(db))
        st.metric("Present Today", today_cnt)
        
        if st.session_state["is_admin"]:
            st.markdown("---")
            threshold = st.slider(
                "Strictness Threshold",
                min_value=0.20, max_value=0.90,
                value=Config.DEFAULT_THRESHOLD, step=0.05,
            )
        else:
            threshold = Config.DEFAULT_THRESHOLD
            
        st.markdown("<br><br><hr>", unsafe_allow_html=True)
        st.markdown("#### Admin Portal")
        pwd = st.text_input("Access Key", type="password", key="admin_pwd_input")
        
        if pwd == Config.ADMIN_PASSWORD:
            if not st.session_state["is_admin"]:
                st.session_state["is_admin"] = True
                st.rerun()
            st.success("Admin mode unlocked.")
        elif pwd != "":
            if st.session_state["is_admin"]:
                st.session_state["is_admin"] = False
                st.rerun()
            st.error("Invalid access key.")
        else:
            if st.session_state["is_admin"]:
                st.session_state["is_admin"] = False
                st.rerun()

    if "deepface_model" not in st.session_state:
        with st.spinner("Initializing AI Models..."):
            st.session_state.deepface_model = load_deepface()
    
    DeepFace = st.session_state.deepface_model

    hero_banner()

    if "Registration" in page:
        page_register(DeepFace, st.session_state["is_admin"])
    elif "Scan" in page:
        page_attendance(DeepFace, threshold, st.session_state["is_admin"])
    elif "Logs" in page:
        page_records()
    elif "Configuration" in page:
        page_settings(threshold, st.session_state["is_admin"])

if __name__ == "__main__":
    main()